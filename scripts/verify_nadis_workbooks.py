#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import sys
from pathlib import Path

from openpyxl import load_workbook

VACCINATION_SHA256 = "458e7af8b47c491f5245f5fc6cc8bbe754bbc23ab63829e88bb2083b813c05ba"
OUTBREAK_SHA256 = "8ea90b4b5c30a66029186905e9aab846bf897f40121d5ec7d7d69acf2964db94"

VACCINATION_HEADERS = (
    "PK",
    "Code",
    "Country * ",
    "Admin Division (Level 1) * ",
    "Admin Division (Level 2)",
    "Year",
    "Month",
    "Reason for the vaccination * ",
    "Species * ",
    "Disease * ",
    "Number of animals vaccinated for the species selected * ",
    "Name of Vaccine",
    "Type of Vaccine",
    "Source of Vaccine",
    "Batch number",
    "Vaccine tested at PANVAC",
)
VACCINATION_IDS = (None, None, 4650, 4651, 3433, None, None, 3434, 3435, 3436, 3437, 3438, 3439, 3440, 3441, 3442)
VACCINATION_VALIDATIONS = {
    ("H5:H239", "fd_3434_reason_for_the_vaccination"),
    ("I5:I239", "fd_3435_species"),
    ("C5:C239", "fd_4650_country"),
    ("P6:P15 P17:P239", "fd_3442_vaccine_tested_at_panvac"),
    ("D5:D239", "admin_division_level_1_4651"),
    ("E5:G239", "admin_division_level_2_3433"),
    ("J5:J239", "fd_3436_disease"),
    ("M5:M239", "fd_3439_type_of_vaccine"),
}

OUTBREAK_SHEETS = (
    "Outbreaks",
    "Animals affected",
    "Bases of Diagnosis",
    "Disease Control Measures",
    "Locations",
)
OUTBREAK_HEADERS = {
    "Outbreaks": (
        "PK", "Code", "Country", "Admin Level 1", "Year * ", "Month", "Disease * ", "Serotype",
        "New or Follow up outbreak", "Number of New outbreaks * ", "Total number of outbreaks",
        "Date of start of outbreak", "Date reported to Vet", "Date investigated * ",
        "Date of final diagnosis * ", "Source of infection", "Outbreak status",
    ),
    "Animals affected": (
        "PK", "parent", "Species * ", "Age Group", "Sex", "Number susceptible", "Number of cases * ",
        "Number of death", "Number slaughtered", "Number destroyed", "Number vaccinated around the outbreak",
    ),
    "Bases of Diagnosis": ("PK", "parent", "Basis of diagnosis"),
    "Disease Control Measures": ("PK", "parent", "Disease control measure * ", "Flag * "),
    "Locations": ("PK", "parent", "Name of locality * ", "Epidemiological unit type", "Production system", "Location coordinate", None),
}
OUTBREAK_IDS = {
    "Outbreaks": (None, None, None, 5518, 3210, 3211, 3212, 3213, 3214, 3220, 3221, 3215, 3713, 3216, 3219, 3222, 3223),
    "Animals affected": (None, None, 4579, 4605, 4581, 4582, 4583, 4585, 4586, 4587, 4588),
    "Bases of Diagnosis": (None, None, 1359),
    "Disease Control Measures": (None, None, 1394, 4606),
    "Locations": (None, None, 1344, 1343, 1346, 1345, None),
}
OUTBREAK_HIDDEN_COLUMNS = {
    "Outbreaks": {"A", "CA"},
    "Animals affected": {"A", "CH"},
    "Bases of Diagnosis": {"A", "CK"},
    "Disease Control Measures": {"A", "CL"},
    "Locations": {"A", "CN"},
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def row_values(sheet, row: int, width: int) -> tuple:
    return tuple(sheet.cell(row=row, column=column).value for column in range(1, width + 1))


def validation_pairs(sheet) -> set[tuple[str, str]]:
    return {(str(item.sqref), str(item.formula1 or "")) for item in sheet.data_validations.dataValidation}


def hidden_columns(sheet) -> set[str]:
    return {key for key, dimension in sheet.column_dimensions.items() if dimension.hidden}


def verify_reference_vaccination(path: Path) -> list[str]:
    errors = []
    if sha256(path) != VACCINATION_SHA256:
        errors.append(f"Vaccination reference hash mismatch: {sha256(path)}")
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != ["Vaccinations"]:
        errors.append(f"Vaccination sheets differ: {workbook.sheetnames}")
        return errors
    sheet = workbook["Vaccinations"]
    if sheet["B1"].value != "Monthly Vaccination Report":
        errors.append("Vaccination B1 title differs")
    if row_values(sheet, 2, 16) != VACCINATION_IDS:
        errors.append("Vaccination field-ID row differs")
    if row_values(sheet, 3, 16) != VACCINATION_HEADERS:
        errors.append("Vaccination header row differs")
    if sheet["B4"].value != 1 or sheet["O4"].value != "u":
        errors.append("Vaccination row-4 metadata differs")
    if not {"A", "CD"}.issubset(hidden_columns(sheet)):
        errors.append(f"Vaccination hidden columns differ: {sorted(hidden_columns(sheet))}")
    if 2 not in {index for index, dimension in sheet.row_dimensions.items() if dimension.hidden}:
        errors.append("Vaccination row 2 is not hidden")
    if validation_pairs(sheet) != VACCINATION_VALIDATIONS:
        errors.append(f"Vaccination validations differ: {sorted(validation_pairs(sheet))}")
    names = {name.name: name.attr_text for name in workbook.defined_names.values()}
    for name, target in {
        "admin_division_level_1_4651": "Vaccinations!$CD$1:$CD$851",
        "admin_division_level_2_3433": "Vaccinations!$CE$1:$CE$670",
    }.items():
        if names.get(name) != target:
            errors.append(f"Vaccination defined name {name} differs: {names.get(name)!r}")
    return errors


def verify_reference_outbreak(path: Path) -> list[str]:
    errors = []
    if sha256(path) != OUTBREAK_SHA256:
        errors.append(f"Outbreak reference hash mismatch: {sha256(path)}")
    workbook = load_workbook(path, data_only=False)
    if tuple(workbook.sheetnames) != OUTBREAK_SHEETS:
        errors.append(f"Outbreak sheets differ: {workbook.sheetnames}")
        return errors
    for name in OUTBREAK_SHEETS:
        sheet = workbook[name]
        width = len(OUTBREAK_HEADERS[name])
        if row_values(sheet, 2, width) != OUTBREAK_IDS[name]:
            errors.append(f"{name}: field-ID row differs")
        if row_values(sheet, 3, width) != OUTBREAK_HEADERS[name]:
            errors.append(f"{name}: header row differs")
        if not OUTBREAK_HIDDEN_COLUMNS[name].issubset(hidden_columns(sheet)):
            errors.append(f"{name}: hidden columns differ: {sorted(hidden_columns(sheet))}")
        if 2 not in {index for index, dimension in sheet.row_dimensions.items() if dimension.hidden}:
            errors.append(f"{name}: row 2 is not hidden")
    for coordinate in ("L4", "M4", "N4", "O4"):
        if workbook["Outbreaks"][coordinate].value != "(dd/mm/yyyy)":
            errors.append(f"Outbreaks {coordinate} date hint differs")
    if workbook["Locations"]["F4"].value != "Latitude" or workbook["Locations"]["G4"].value != "Longitude":
        errors.append("Locations coordinate subheaders differ")
    return errors


def verify_generated_vaccination(path: Path) -> list[str]:
    errors = []
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != ["Vaccinations"]:
        errors.append(f"Generated vaccination sheets differ: {workbook.sheetnames}")
        return errors
    sheet = workbook["Vaccinations"]
    if row_values(sheet, 2, 16) != VACCINATION_IDS:
        errors.append("Generated vaccination field-ID row differs")
    if row_values(sheet, 3, 16) != VACCINATION_HEADERS:
        errors.append("Generated vaccination header row differs")
    if sheet["B1"].value != "Monthly Vaccination Report":
        errors.append("Generated vaccination title differs")
    for row in range(5, min(sheet.max_row, 239) + 1):
        if sheet.cell(row=row, column=3).value and not sheet.cell(row=row, column=11).value:
            errors.append(f"Generated vaccination row {row} has data but no vaccinated count")
    return errors


def verify_generated_outbreak(path: Path) -> list[str]:
    errors = []
    workbook = load_workbook(path, data_only=False)
    if tuple(workbook.sheetnames) != OUTBREAK_SHEETS:
        errors.append(f"Generated outbreak sheets differ: {workbook.sheetnames}")
        return errors
    parent_codes = {
        workbook["Outbreaks"].cell(row=row, column=2).value
        for row in range(5, workbook["Outbreaks"].max_row + 1)
        if workbook["Outbreaks"].cell(row=row, column=2).value
    }
    for name in OUTBREAK_SHEETS:
        sheet = workbook[name]
        width = len(OUTBREAK_HEADERS[name])
        if row_values(sheet, 2, width) != OUTBREAK_IDS[name]:
            errors.append(f"Generated {name}: field-ID row differs")
        if row_values(sheet, 3, width) != OUTBREAK_HEADERS[name]:
            errors.append(f"Generated {name}: header row differs")
    for name in OUTBREAK_SHEETS[1:]:
        sheet = workbook[name]
        for row in range(5, sheet.max_row + 1):
            parent = sheet.cell(row=row, column=2).value
            if parent and parent not in parent_codes:
                errors.append(f"Generated {name} row {row} references missing parent {parent}")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description="Verify supplied and VetEdge-generated VCN/NADIS workbooks.")
    parser.add_argument("--vaccination-reference", type=Path)
    parser.add_argument("--outbreak-reference", type=Path)
    parser.add_argument("--vaccination-generated", type=Path)
    parser.add_argument("--outbreak-generated", type=Path)
    args = parser.parse_args()

    checks = []
    if args.vaccination_reference:
        checks.append(("vaccination reference", verify_reference_vaccination(args.vaccination_reference)))
    if args.outbreak_reference:
        checks.append(("outbreak reference", verify_reference_outbreak(args.outbreak_reference)))
    if args.vaccination_generated:
        checks.append(("generated vaccination", verify_generated_vaccination(args.vaccination_generated)))
    if args.outbreak_generated:
        checks.append(("generated outbreak", verify_generated_outbreak(args.outbreak_generated)))
    if not checks:
        parser.error("Provide at least one workbook path to verify.")

    failed = False
    for label, errors in checks:
        if errors:
            failed = True
            print(f"FAIL: {label}")
            for error in errors:
                print(f"  - {error}")
        else:
            print(f"PASS: {label}")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
