from __future__ import annotations

from collections import defaultdict
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import cint, cstr, get_datetime

from vetedge.install.custom_fields import (
    BRANCH_NADIS_ADMIN_LEVEL_1_FIELD,
    BRANCH_NADIS_ADMIN_LEVEL_2_FIELD,
)
from vetedge.services.nadis_reporting import (
    PATIENT_DOCTYPE,
    VACCINATION_DOCTYPE,
    _filters,
    _query_filters,
    _require_permissions,
)
from vetedge.services.nadis_templates import (
    PANVAC_VALUES,
    VACCINATION_DATA_START_ROW,
    VACCINATION_FIELD_IDS,
    VACCINATION_HEADERS,
    VACCINATION_REASONS,
    VACCINATION_SHEET,
    VACCINATION_TEMPLATE_FILENAME,
    VACCINATION_TEMPLATE_SHA256,
    VACCINATION_TITLE,
    VACCINATION_VACCINE_TYPES,
)

VACCINE_DOCTYPE = "Veterinary Vaccine"
SPECIES_DOCTYPE = "Veterinary Species"
BRANCH_DOCTYPE = "Branch"
EXPORT_FETCH_PAGE = 500
MAX_SOURCE_ROWS = 20_000
MAX_TEMPLATE_DATA_ROWS = 235  # official template validates rows 5:239


def _require_export_permissions() -> None:
    _require_permissions()
    for doctype, label in (
        (VACCINE_DOCTYPE, _("vaccines")),
        (SPECIES_DOCTYPE, _("species mappings")),
        (BRANCH_DOCTYPE, _("branches")),
    ):
        if not frappe.has_permission(doctype, "read"):
            frappe.throw(
                _("You do not have permission to read {0} required for NADIS export.").format(label),
                frappe.PermissionError,
            )


def _fetch_source_rows(query_filters: dict) -> list[dict]:
    total = cint(frappe.db.count(VACCINATION_DOCTYPE, filters=query_filters))
    if total > MAX_SOURCE_ROWS:
        frappe.throw(
            _("The selected period contains {0} vaccination records. Narrow the filters below {1} records before generating the official workbook.").format(total, MAX_SOURCE_ROWS),
            frappe.ValidationError,
        )

    rows: list[dict] = []
    start = 0
    while start < total:
        page = frappe.get_list(
            VACCINATION_DOCTYPE,
            filters=query_filters,
            fields=[
                "name",
                "patient",
                "status",
                "service_branch",
                "company",
                "vaccine",
                "administered_on",
                "vaccination_reason",
                "batch_no",
            ],
            order_by="administered_on asc, name asc",
            start=start,
            page_length=min(EXPORT_FETCH_PAGE, total - start),
        )
        if not page:
            break
        rows.extend(page)
        start += len(page)
    return rows


def _chunked(values: list[str], size: int = 500):
    for start in range(0, len(values), size):
        yield values[start : start + size]


def _patient_species_map(patient_names: list[str]) -> dict[str, str]:
    names = sorted({cstr(value).strip() for value in patient_names if cstr(value).strip()})
    result: dict[str, str] = {}
    for chunk in _chunked(names):
        for row in frappe.get_list(
            PATIENT_DOCTYPE,
            filters={"name": ["in", chunk]},
            fields=["name", "species"],
            page_length=len(chunk),
        ):
            result[row.get("name")] = row.get("species")
    return result


def _species_mapping(species_names: list[str]) -> dict[str, str]:
    names = sorted({cstr(value).strip() for value in species_names if cstr(value).strip()})
    if not names:
        return {}
    rows = frappe.get_list(
        SPECIES_DOCTYPE,
        filters={"name": ["in", names]},
        fields=["name", "nadis_species"],
        page_length=len(names),
    )
    return {row.get("name"): cstr(row.get("nadis_species")).strip() for row in rows}


def _vaccine_mapping(vaccine_names: list[str]) -> dict[str, dict]:
    names = sorted({cstr(value).strip() for value in vaccine_names if cstr(value).strip()})
    if not names:
        return {}
    rows = frappe.get_list(
        VACCINE_DOCTYPE,
        filters={"name": ["in", names]},
        fields=[
            "name",
            "vaccine_name",
            "nadis_disease",
            "nadis_vaccine_type",
            "nadis_source_of_vaccine",
            "nadis_panvac_tested",
        ],
        page_length=len(names),
    )
    return {row.get("name"): row for row in rows}


def _branch_mapping(branch_names: list[str]) -> dict[str, dict]:
    names = sorted({cstr(value).strip() for value in branch_names if cstr(value).strip()})
    if not names:
        return {}
    meta = frappe.get_meta(BRANCH_DOCTYPE)
    fields = ["name"]
    for fieldname in (BRANCH_NADIS_ADMIN_LEVEL_1_FIELD, BRANCH_NADIS_ADMIN_LEVEL_2_FIELD):
        if meta.has_field(fieldname):
            fields.append(fieldname)
    rows = frappe.get_list(
        BRANCH_DOCTYPE,
        filters={"name": ["in", names]},
        fields=fields,
        page_length=len(names),
    )
    return {row.get("name"): row for row in rows}


def _month_name(value) -> str:
    return get_datetime(value).strftime("%B")


def _validation_result(source_rows: list[dict]) -> dict:
    patients = _patient_species_map([row.get("patient") for row in source_rows])
    species_map = _species_mapping(list(patients.values()))
    vaccine_map = _vaccine_mapping([row.get("vaccine") for row in source_rows])
    branch_map = _branch_mapping([row.get("service_branch") for row in source_rows])

    errors: list[dict] = []
    warnings: list[dict] = []
    prepared: list[dict] = []

    for row in source_rows:
        record = row.get("name")
        patient = cstr(row.get("patient")).strip()
        vaccine = vaccine_map.get(row.get("vaccine")) or {}
        branch = branch_map.get(row.get("service_branch")) or {}
        source_species = patients.get(patient)
        nadis_species = cstr(species_map.get(source_species)).strip()
        administered_on = row.get("administered_on")
        reason = cstr(row.get("vaccination_reason")).strip()
        disease = cstr(vaccine.get("nadis_disease")).strip()
        admin_level_1 = cstr(branch.get(BRANCH_NADIS_ADMIN_LEVEL_1_FIELD)).strip()
        admin_level_2 = cstr(branch.get(BRANCH_NADIS_ADMIN_LEVEL_2_FIELD)).strip()
        vaccine_type = cstr(vaccine.get("nadis_vaccine_type")).strip()
        panvac = cstr(vaccine.get("nadis_panvac_tested")).strip()

        missing = []
        if not patient:
            missing.append("Patient")
        if not administered_on:
            missing.append("Administered On")
        if not row.get("service_branch"):
            missing.append("Service Branch")
        if not admin_level_1:
            missing.append("Branch NADIS State / Admin Level 1")
        if not reason:
            missing.append("Reason for Vaccination")
        if not nadis_species:
            missing.append("Species NADIS mapping")
        if not disease:
            missing.append("Vaccine NADIS Disease")

        if missing:
            errors.append({"record": record, "message": _("Missing required NADIS data: {0}").format(", ".join(missing))})
            continue
        if reason not in VACCINATION_REASONS:
            errors.append({"record": record, "message": _("Reason for Vaccination is not one of the two values in the supplied NADIS template.")})
            continue
        if vaccine_type and vaccine_type not in VACCINATION_VACCINE_TYPES:
            errors.append({"record": record, "message": _("Vaccine Type does not match the supplied NADIS template values.")})
            continue
        if panvac and panvac not in PANVAC_VALUES:
            errors.append({"record": record, "message": _("PANVAC value must be Yes or No.")})
            continue
        if admin_level_1 and not admin_level_1.endswith(", Nigeria"):
            warnings.append({"record": record, "message": _("Admin Level 1 does not use the supplied workbook's '..., Nigeria' wording.")})
        if admin_level_2 and not admin_level_2.endswith(", Nigeria"):
            warnings.append({"record": record, "message": _("Admin Level 2 does not use the supplied workbook's '..., Nigeria' wording.")})

        prepared.append(
            {
                "record": record,
                "patient": patient,
                "country": "Nigeria",
                "admin_level_1": admin_level_1,
                "admin_level_2": admin_level_2,
                "year": get_datetime(administered_on).year,
                "month": _month_name(administered_on),
                "reason": reason,
                "species": nadis_species,
                "disease": disease,
                "vaccine_name": cstr(vaccine.get("vaccine_name") or row.get("vaccine")).strip(),
                "vaccine_type": vaccine_type,
                "source_of_vaccine": cstr(vaccine.get("nadis_source_of_vaccine")).strip(),
                "batch_no": cstr(row.get("batch_no")).strip(),
                "panvac": panvac,
            }
        )

    return {"prepared": prepared, "errors": errors, "warnings": warnings}


def _group_key(row: dict) -> tuple:
    return (
        row["country"],
        row["admin_level_1"],
        row["admin_level_2"],
        row["year"],
        row["month"],
        row["reason"],
        row["species"],
        row["disease"],
        row["vaccine_name"],
        row["vaccine_type"],
        row["source_of_vaccine"],
        row["batch_no"],
        row["panvac"],
    )


def _aggregate(prepared: list[dict]) -> tuple[list[dict], list[dict]]:
    patients_by_group: dict[tuple, set[str]] = defaultdict(set)
    records_by_group: dict[tuple, list[str]] = defaultdict(list)
    for row in prepared:
        key = _group_key(row)
        patients_by_group[key].add(row["patient"])
        records_by_group[key].append(row["record"])

    rows = []
    warnings = []
    for key in sorted(patients_by_group):
        (
            country,
            admin_level_1,
            admin_level_2,
            year,
            month,
            reason,
            species,
            disease,
            vaccine_name,
            vaccine_type,
            source_of_vaccine,
            batch_no,
            panvac,
        ) = key
        distinct_patients = patients_by_group[key]
        source_records = records_by_group[key]
        if len(source_records) > len(distinct_patients):
            warnings.append(
                {
                    "record": None,
                    "message": _("{0} vaccination records in one NADIS grouping represent {1} distinct animals; VetEdge reports the distinct animal count to avoid double-counting.").format(len(source_records), len(distinct_patients)),
                }
            )
        rows.append(
            {
                "country": country,
                "admin_level_1": admin_level_1,
                "admin_level_2": admin_level_2,
                "year": year,
                "month": month,
                "reason": reason,
                "species": species,
                "disease": disease,
                "number_vaccinated": len(distinct_patients),
                "vaccine_name": vaccine_name,
                "vaccine_type": vaccine_type,
                "source_of_vaccine": source_of_vaccine,
                "batch_no": batch_no,
                "panvac": panvac,
            }
        )
    return rows, warnings


def _official_rows(filters: str | dict | None = None) -> dict:
    _require_export_permissions()
    report_filters = _filters(filters)
    query_filters = _query_filters(report_filters)
    query_filters["status"] = "Administered"
    source_rows = _fetch_source_rows(query_filters)
    validation = _validation_result(source_rows)
    aggregated, aggregation_warnings = _aggregate(validation["prepared"])
    validation["warnings"].extend(aggregation_warnings)
    if len(aggregated) > MAX_TEMPLATE_DATA_ROWS:
        validation["errors"].append(
            {
                "record": None,
                "message": _("The official vaccination template supports {0} mapped data rows (rows 5-239); this selection produces {1} grouped rows. Narrow the reporting scope.").format(MAX_TEMPLATE_DATA_ROWS, len(aggregated)),
            }
        )
    return {
        "source_count": len(source_rows),
        "distinct_animal_count": sum(row["number_vaccinated"] for row in aggregated),
        "rows": aggregated,
        "errors": validation["errors"],
        "warnings": validation["warnings"],
        "template_mapping_verified": True,
        "template_filename": VACCINATION_TEMPLATE_FILENAME,
        "template_sha256": VACCINATION_TEMPLATE_SHA256,
        "submission_ready": not validation["errors"] and bool(aggregated),
    }


@frappe.whitelist()
@frappe.read_only()
def validate_nadis_vaccination_export(filters: str | dict | None = None) -> dict:
    result = _official_rows(filters)
    return {
        "source_count": result["source_count"],
        "distinct_animal_count": result["distinct_animal_count"],
        "grouped_row_count": len(result["rows"]),
        "errors": result["errors"][:200],
        "warnings": result["warnings"][:200],
        "error_count": len(result["errors"]),
        "warning_count": len(result["warnings"]),
        "template_mapping_verified": True,
        "template_filename": result["template_filename"],
        "template_sha256": result["template_sha256"],
        "submission_ready": result["submission_ready"],
    }


def _build_workbook(rows: list[dict]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        frappe.throw(_("openpyxl is required to generate the official NADIS workbook."))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = VACCINATION_SHEET
    sheet["B1"] = VACCINATION_TITLE
    sheet["B1"].font = Font(bold=True, size=14)

    for column, field_id in enumerate(VACCINATION_FIELD_IDS, start=1):
        sheet.cell(row=2, column=column, value=field_id)
    sheet.row_dimensions[2].hidden = True
    for column, header in enumerate(VACCINATION_HEADERS, start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Preserve the non-data row-4 markers present in the supplied workbook.
    sheet["B4"] = 1
    sheet["O4"] = "u"
    sheet.column_dimensions["A"].hidden = True
    sheet.column_dimensions["CD"].hidden = True
    widths = {"B": 10, "C": 16, "D": 30, "E": 30, "F": 10, "G": 14, "H": 28, "I": 20, "J": 34, "K": 18, "L": 28, "M": 24, "N": 28, "O": 22, "P": 24}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    reason_validation = DataValidation(type="list", formula1='"Control/Emergency vaccination,Preventive/Routine vaccination"', allow_blank=False)
    vaccine_type_validation = DataValidation(type="list", formula1='"Anti-idiotype vaccines,Conjugate vaccines,DNA vaccines,Inactivated vaccines"', allow_blank=True)
    panvac_validation = DataValidation(type="list", formula1='"No,Yes"', allow_blank=True)
    sheet.add_data_validation(reason_validation)
    sheet.add_data_validation(vaccine_type_validation)
    sheet.add_data_validation(panvac_validation)
    reason_validation.add(f"H{VACCINATION_DATA_START_ROW}:H239")
    vaccine_type_validation.add(f"M{VACCINATION_DATA_START_ROW}:M239")
    panvac_validation.add(f"P{VACCINATION_DATA_START_ROW}:P239")

    for offset, row in enumerate(rows):
        excel_row = VACCINATION_DATA_START_ROW + offset
        values = (
            None,
            offset + 1,
            row["country"],
            row["admin_level_1"],
            row["admin_level_2"],
            row["year"],
            row["month"],
            row["reason"],
            row["species"],
            row["disease"],
            row["number_vaccinated"],
            row["vaccine_name"],
            row["vaccine_type"],
            row["source_of_vaccine"],
            row["batch_no"],
            row["panvac"],
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row=excel_row, column=column, value=value)

    workbook.properties.title = VACCINATION_TITLE
    workbook.properties.subject = "VCN / NADIS regulatory vaccination report generated by VetEdge"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _export_filename(report_filters: dict) -> str:
    from_date = cstr(report_filters.get("from_date")).strip()
    to_date = cstr(report_filters.get("to_date")).strip()
    suffix = ""
    if from_date or to_date:
        suffix = "_" + "_to_".join(value for value in (from_date, to_date) if value)
    return f"NADIS_Monthly_Vaccination_Report{suffix}.xlsx"


@frappe.whitelist()
def download_nadis_vaccination_workbook(filters: str | dict | None = None):
    """Generate the mapped vaccination workbook or fail before download.

    This endpoint is intentionally separate from the generic EdgeSuite XLSX
    exporter so presentation-export changes cannot alter the regulatory schema.
    """
    result = _official_rows(filters)
    if not result["submission_ready"]:
        messages = [item.get("message") for item in result["errors"][:10]]
        frappe.throw(
            _("NADIS export is blocked until required regulatory data is complete.{0}").format("\n" + "\n".join(messages) if messages else ""),
            frappe.ValidationError,
        )

    report_filters = _filters(filters)
    payload = _build_workbook(result["rows"])
    frappe.local.response.filename = _export_filename(report_filters)
    frappe.local.response.filecontent = payload
    frappe.local.response.type = "binary"
    frappe.local.response.display_content_as = "attachment"
